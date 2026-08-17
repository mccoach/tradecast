import tkinter as tk
from tkinter import filedialog, messagebox
import os
import re
import json
import chardet
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  常量 & 规则
# ─────────────────────────────────────────────

APP_NAME = "TradeCast"
DEFAULT_OUT_DIR = r"E:\陈谨\Desktop"
DEFAULT_OPEN_DIR = r"E:\陈谨\Desktop"
CONFIG_FILENAME = "TradeCast_config.json"
_EXCLUDED = object()  # 主动过滤哨兵

# 前两位 → (市场, 标的类型)
CODE_PREFIX_MAP = {
    "60": ("SH", "股票"),
    "68": ("SH", "股票"),
    "90": ("SH", "股票"),
    "51": ("SH", "基金"),
    "58": ("SH", "基金"),
    "00": ("SZ", "股票"),
    "30": ("SZ", "股票"),
    "15": ("SZ", "基金"),
    "16": ("SZ", "基金"),
}

TYPE1_REQUIRED = {
    "成交日期", "成交时间", "证券代码", "证券名称", "委托类别", "成交价格", "成交数量", "发生金额", "股东代码"
}

TYPE2_REQUIRED = {"成交日期", "成交时间", "委托类别", "发生金额", "摘要", "资金帐号"}

TYPE2_EXCLUDE_PATTERNS = [
    r'[A-Z0-9]{4,}[买卖]',
    r'基金申购',
    r'红利到[帐账]',
    r'确认金额',
]

# ─────────────────────────────────────────────
#  配置读写
# ─────────────────────────────────────────────


def _config_path():
    base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, CONFIG_FILENAME)


def load_config():
    try:
        with open(_config_path(), "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_config(data):
    try:
        path = _config_path()
        existing = load_config()
        existing.update(data)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(existing, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# ─────────────────────────────────────────────
#  工具函数
# ─────────────────────────────────────────────


def detect_encoding(raw_bytes):
    result = chardet.detect(raw_bytes)
    enc = result.get("encoding") or "utf-8"
    if enc.lower() in ("gb2312", "gbk", "gb18030"):
        enc = "gb18030"
    return enc


def lookup_code(code):
    prefix = str(code).strip().zfill(6)[:2]
    return CODE_PREFIX_MAP.get(prefix, (None, None))


def fmt_date(val):
    s = str(val).strip().replace("-", "").replace("/", "")
    if len(s) == 8 and s.isdigit():
        return f"{s[:4]}-{s[4:6]}-{s[6:]}"
    return s


def extract_bank_name(summary):
    m = re.search(r'[\u4e00-\u9fa5]{2,4}存管', summary)
    return m.group(0) if m else "存管"


# ─────────────────────────────────────────────
#  文本解析
# ─────────────────────────────────────────────


def parse_text(text):
    lines = [l for l in text.splitlines() if l.strip()]
    rows = []
    for line in lines:
        if "\t" in line:
            parts = line.split("\t")
        else:
            parts = re.split(r" {2,}", line)
        rows.append([p.strip() for p in parts if p.strip()])
    return rows


def find_header_row(rows, required_fields):
    for i, row in enumerate(rows):
        if required_fields.issubset(set(row)):
            col_map = {field: row.index(field) for field in required_fields}
            return i, col_map
    return None, None


def build_df_from_rows(rows, header_idx, col_map):
    records = []
    for row in rows[header_idx + 1:]:
        if not any(row):
            continue
        record = {
            field: (row[idx] if idx < len(row) else "")
            for field, idx in col_map.items()
        }
        records.append(record)
    return records


def detect_and_parse(text):
    rows = parse_text(text)
    hi, col_map = find_header_row(rows, TYPE1_REQUIRED)
    if hi is not None:
        return 1, build_df_from_rows(rows, hi, col_map), col_map
    hi, col_map = find_header_row(rows, TYPE2_REQUIRED)
    if hi is not None:
        return 2, build_df_from_rows(rows, hi, col_map), col_map
    return 0, [], {}


# ─────────────────────────────────────────────
#  智能文件名推断
# ─────────────────────────────────────────────


def infer_filename(text):
    rtype, records, _ = detect_and_parse(text)
    if not records:
        return "交易记录导出"
    dates = []
    for r in records:
        d = str(r.get("成交日期", "")).strip().replace("-", "").replace("/", "")
        if len(d) == 8 and d.isdigit():
            dates.append(d)
    if dates:
        d_min, d_max = min(dates), max(dates)
        date_range = f"{d_min}-{d_max}" if d_min != d_max else d_min
    else:
        date_range = "未知日期"
    label = "转账记录" if rtype == 2 else "交易记录"
    return f"{date_range}{label}"


# ─────────────────────────────────────────────
#  第一类处理
# ─────────────────────────────────────────────


def process_type1(records):
    unknown_ops, unknown_codes, results = set(), set(), []
    for row in records:
        try:
            amt = float(str(row.get("发生金额", "0")).replace(",", ""))
        except Exception:
            amt = 0.0
        if amt == 0:
            continue
        op_raw = str(row.get("委托类别", "")).strip()
        code_raw = str(row.get("证券代码", "")).strip().zfill(6)
        name = str(row.get("证券名称", "")).strip()
        date = fmt_date(row.get("成交日期", ""))
        time_ = str(row.get("成交时间", "")).strip()
        price = str(row.get("成交价格", "")).strip()
        qty_raw = str(row.get("成交数量", "0")).replace(",", "").strip()
        account = str(row.get("股东代码", "")).strip()

        market, asset_type = lookup_code(code_raw)
        if market is None:
            unknown_codes.add(code_raw)
            market, asset_type = "??", "未知"
        full_code = f"{market}{code_raw}"

        if op_raw == "买入": op_key = f"{asset_type}买入"
        elif op_raw == "卖出": op_key = f"{asset_type}卖出"
        elif op_raw == "基金申购": op_key = "基金申购"
        elif op_raw == "红利": op_key = "股息红利"
        else:
            unknown_ops.add(op_raw)
            op_key = op_raw

        if op_key == "股息红利":
            price = str(abs(amt))
            qty = -1
        else:
            try:
                qty = int(float(qty_raw))
            except Exception:
                qty = 0
            if "卖出" in op_key:
                qty = -abs(qty)

        results.append({
            "成交日期": date,
            "成交时间": time_,
            "证券代码": full_code,
            "证券名称": name,
            "操作_公式": op_key,
            "成交均价": price,
            "成交数量": qty,
            "股东帐户": account,
            "名称是公式": False,
        })
    return results, unknown_ops, unknown_codes


# ─────────────────────────────────────────────
#  第二类处理
# ─────────────────────────────────────────────


def classify_type2(summary):
    for pat in TYPE2_EXCLUDE_PATTERNS:
        if re.search(pat, summary):
            return _EXCLUDED  # 主动过滤
    if re.search(r'结息|利息', summary):
        return ("利息归本", "")
    if "存管" in summary:
        if "转入" in summary:
            return ("银证转入", extract_bank_name(summary))
        if "转出" in summary:
            return ("银证转出", extract_bank_name(summary))
    return None  # 真正无法识别


def process_type2(records):
    unknown_summaries, excluded_summaries, results = [], [], []
    for row in records:
        summary = str(row.get("摘要", "")).strip()
        amount_raw = str(row.get("发生金额", "0")).replace(",", "").strip()
        date = fmt_date(row.get("成交日期", ""))
        time_ = str(row.get("成交时间", "")).strip()
        try:
            amount = float(amount_raw)
        except Exception:
            amount = 0.0
        classified = classify_type2(summary)
        if classified is _EXCLUDED:
            if summary:
                excluded_summaries.append(summary)
            continue
        if classified is None:
            if summary:
                unknown_summaries.append(summary)
            continue
        op_key, _ = classified
        if op_key in ("银证转入", "银证转出"):
            name_val = 'INDEX($BL:$BL,MATCH(INDEX($Y:$Y,ROW()),$BK:$BK,0))'
        elif op_key == "利息归本":
            name_val = "留空不要复制"
        else:
            name_val = ""
        results.append({
            "成交日期": date,
            "成交时间": time_,
            "证券代码": "",
            "证券名称": name_val,
            "操作_公式": op_key,
            "成交均价": abs(amount),
            "成交数量": -1 if amount > 0 else 1,
            "股东帐户": "",
            "名称是公式": False,
        })
    return results, unknown_summaries, excluded_summaries


# ─────────────────────────────────────────────
#  写出 xlsx
# ─────────────────────────────────────────────

OUTPUT_HEADERS = ["成交日期", "成交时间", "证券代码", "证券名称", "操作", "成交均价", "成交数量", "股东帐户"]
COL_WIDTHS = [12, 10, 14, 24, 12, 12, 10, 14]
LEFT_COLS = {3, 4, 8}


def write_xlsx(all_rows, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "交易记录"
    header_fill = PatternFill("solid", fgColor="2B2B2B")
    header_font = Font(color="FFFFFF", bold=True, name="微软雅黑", size=10)
    even_fill = PatternFill("solid", fgColor="F5F5F5")
    odd_fill = PatternFill("solid", fgColor="FFFFFF")
    border_side = Side(style="thin", color="DDDDDD")
    cell_border = Border(left=border_side,
                         right=border_side,
                         top=border_side,
                         bottom=border_side)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    for col, h in enumerate(OUTPUT_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center_align
        c.border = cell_border

    for row_idx, r in enumerate(all_rows, 2):
        fill = even_fill if row_idx % 2 == 0 else odd_fill

        def wc(col, value, num_fmt=None):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = value
            cell.fill = fill
            cell.border = cell_border
            cell.alignment = left_align if col in LEFT_COLS else center_align
            if num_fmt:
                cell.number_format = num_fmt

        wc(1, r["成交日期"], num_fmt="YYYY-MM-DD")
        wc(2, r["成交时间"], num_fmt="HH:MM:SS")
        wc(3, r["证券代码"])
        wc(4, r["证券名称"])
        wc(5, r["操作_公式"])
        wc(6, r["成交均价"], num_fmt="0.0000")
        wc(7, r["成交数量"], num_fmt="0")
        wc(8, r["股东帐户"], num_fmt="@")

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(out_path)


# ─────────────────────────────────────────────
#  核心处理入口
# ─────────────────────────────────────────────


def process_text(text, out_path):
    rtype, records, _ = detect_and_parse(text)
    if rtype == 0:
        return False, "无法识别记录类型，请检查内容格式"

    if rtype == 1:
        all_rows, unknown_ops, unknown_codes = process_type1(records)
        unknown_summaries, excluded_summaries = [], []
    else:
        all_rows, unknown_summaries, excluded_summaries = process_type2(
            records)
        unknown_ops = unknown_codes = set()

    warnings = []
    if unknown_ops:
        warnings.append(f"⚠️ 发现未知委托类别，请补充规则：\n{', '.join(unknown_ops)}")
    if unknown_codes:
        warnings.append(f"⚠️ 以下代码无法判断市场，请手动确认：\n{', '.join(unknown_codes)}")
    if excluded_summaries:
        uniq = list(dict.fromkeys(excluded_summaries))[:20]
        warnings.append("🔕 以下摘要已主动过滤（重复交易流水，正常）：\n" + "\n".join(uniq))
    if unknown_summaries:
        uniq = list(dict.fromkeys(unknown_summaries))[:20]
        warnings.append("⚠️ 以下摘要无法分类（已跳过，请检查规则）：\n" + "\n".join(uniq))

    write_xlsx(all_rows, out_path)
    msg = f"✅ 导出成功！共 {len(all_rows)} 条\n路径：{out_path}"
    if warnings:
        msg += "\n\n" + "\n\n".join(warnings)
    return True, msg


# ─────────────────────────────────────────────
#  Tk 界面
# ─────────────────────────────────────────────

BG = "#F0F0F0"
PANEL = "#FFFFFF"
DARK = "#2B2B2B"
MID = "#666666"
GHOST = "#EEEEEE"
BORDER = "#DDDDDD"


class App(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title(APP_NAME)
        self.geometry("960x700")
        self.resizable(True, True)
        self.configure(bg=BG)
        self._debounce_id = None
        self._suppress_name_save = False  # 程序自动改文件名时置 True，跳过保存
        self._cfg = load_config()
        self._build_ui()
        self._apply_config()

    # ── 配置应用 ─────────────────────────────

    def _apply_config(self):
        out_dir = self._cfg.get("out_dir") or DEFAULT_OUT_DIR
        self._dir_var.set(out_dir)
        self._auto_open_var.set(bool(self._cfg.get("auto_open", True)))
        self._update_preview()

    # ── UI 构建 ──────────────────────────────

    def _build_ui(self):
        self._build_titlebar()
        self._build_body()
        self._build_statusbar()

    def _build_titlebar(self):
        bar = tk.Frame(self, bg=DARK, height=52)
        bar.pack(fill="x")
        bar.pack_propagate(False)
        tk.Label(bar,
                 text=APP_NAME,
                 bg=DARK,
                 fg="#FFFFFF",
                 font=("微软雅黑", 13, "bold")).pack(side="left", padx=20, pady=14)
        tk.Label(bar,
                 text="交易记录格式转换  ·  TXT / CSV / XLS / XLSX",
                 bg=DARK,
                 fg="#999999",
                 font=("微软雅黑", 9)).pack(side="left", padx=4)

    def _build_body(self):
        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=20, pady=16)
        self._build_left(body)
        self._build_right(body)

    def _build_left(self, parent):
        left = tk.Frame(parent, bg=BG)
        left.pack(side="left", fill="both", expand=True)

        label_row = tk.Frame(left, bg=BG)
        label_row.pack(fill="x", pady=(0, 6))
        tk.Label(label_row,
                 text="原始记录",
                 bg=BG,
                 fg=DARK,
                 font=("微软雅黑", 10, "bold")).pack(side="left")
        tk.Label(label_row,
                 text="粘贴内容或从文件读取，支持随时手动修改",
                 bg=BG,
                 fg=MID,
                 font=("微软雅黑", 8)).pack(side="left", padx=8)

        text_card = self._card(left)
        text_card.pack(fill="both", expand=True)
        tf = tk.Frame(text_card, bg=PANEL)
        tf.pack(fill="both", expand=True, padx=1, pady=1)
        sb_y = tk.Scrollbar(tf, orient="vertical")
        sb_y.pack(side="right", fill="y")
        sb_x = tk.Scrollbar(tf, orient="horizontal")
        sb_x.pack(side="bottom", fill="x")
        self._text_input = tk.Text(tf,
                                   yscrollcommand=sb_y.set,
                                   xscrollcommand=sb_x.set,
                                   bg="#FAFAFA",
                                   fg=DARK,
                                   insertbackground=DARK,
                                   font=("Consolas", 9),
                                   relief="flat",
                                   borderwidth=0,
                                   highlightthickness=0,
                                   wrap="none")
        self._text_input.pack(side="left", fill="both", expand=True)
        sb_y.config(command=self._text_input.yview)
        sb_x.config(command=self._text_input.xview)
        self._text_input.bind("<<Modified>>", self._on_text_modified)

        btn_row = tk.Frame(left, bg=BG)
        btn_row.pack(fill="x", pady=(8, 0))
        self._btn(btn_row, "清空内容", self._clear_text,
                  ghost=True).pack(side="left", padx=(0, 8))
        self._btn(btn_row, "从文件读取", self._load_file).pack(side="left")
        self._detect_label = tk.Label(btn_row,
                                      text="",
                                      bg=BG,
                                      fg=MID,
                                      font=("微软雅黑", 8))
        self._detect_label.pack(side="left", padx=12)

    def _build_right(self, parent):
        right = tk.Frame(parent, bg=BG, width=260)
        right.pack(side="right", fill="y", padx=(16, 0))
        right.pack_propagate(False)

        # ── 输出目录 ──
        dir_card = self._card(right)
        dir_card.pack(fill="x")
        tk.Label(dir_card, text="输出目录", bg=PANEL, fg=MID,
                 font=("微软雅黑", 9)).pack(anchor="w", padx=14, pady=(12, 4))
        dir_row = tk.Frame(dir_card, bg=PANEL)
        dir_row.pack(fill="x", padx=14, pady=(0, 12))
        self._dir_var = tk.StringVar()
        tk.Entry(dir_row,
                 textvariable=self._dir_var,
                 bg="#FAFAFA",
                 fg=DARK,
                 font=("微软雅黑", 9),
                 relief="flat",
                 borderwidth=0,
                 highlightthickness=1,
                 highlightbackground=BORDER).pack(side="left",
                                                  fill="x",
                                                  expand=True,
                                                  ipady=5)
        self._btn(dir_row, "浏览", self._browse_dir,
                  ghost=True).pack(side="left", padx=(6, 0))
        # 用户手动编辑目录框 → 保存
        self._dir_var.trace_add("write", self._on_dir_changed)

        # ── 文件名 ──
        name_card = self._card(right)
        name_card.pack(fill="x", pady=(10, 0))
        tk.Label(name_card,
                 text="文件名（自动生成，可手动修改）",
                 bg=PANEL,
                 fg=MID,
                 font=("微软雅黑", 9)).pack(anchor="w", padx=14, pady=(12, 4))
        name_row = tk.Frame(name_card, bg=PANEL)
        name_row.pack(fill="x", padx=14, pady=(0, 4))
        self._name_var = tk.StringVar(value="交易记录导出")
        tk.Entry(name_row,
                 textvariable=self._name_var,
                 bg="#FAFAFA",
                 fg=DARK,
                 font=("微软雅黑", 9),
                 relief="flat",
                 borderwidth=0,
                 highlightthickness=1,
                 highlightbackground=BORDER).pack(fill="x", ipady=5)

        # .xlsx 标签 + 自动打开勾选框（同行）
        ext_row = tk.Frame(name_card, bg=PANEL)
        ext_row.pack(fill="x", padx=14, pady=(0, 10))
        self._auto_open_var = tk.BooleanVar(value=True)
        tk.Checkbutton(ext_row,
                       text="自动打开成果文件",
                       variable=self._auto_open_var,
                       bg=PANEL,
                       fg=MID,
                       font=("微软雅黑", 8),
                       activebackground=PANEL,
                       command=self._on_auto_open_changed).pack(side="left")
        tk.Label(ext_row,
                 text=".xlsx",
                 bg=PANEL,
                 fg="#AAAAAA",
                 font=("微软雅黑", 8)).pack(side="right")

        # ── 完整路径预览 ──
        preview_card = self._card(right)
        preview_card.pack(fill="x", pady=(10, 0))
        tk.Label(preview_card,
                 text="完整路径预览",
                 bg=PANEL,
                 fg=MID,
                 font=("微软雅黑", 9)).pack(anchor="w", padx=14, pady=(10, 4))
        self._preview_label = tk.Label(preview_card,
                                       text="",
                                       bg=PANEL,
                                       fg="#888888",
                                       font=("微软雅黑", 8),
                                       wraplength=220,
                                       justify="left")
        self._preview_label.pack(anchor="w", padx=14, pady=(0, 10))

        self._name_var.trace_add("write", lambda *_: self._update_preview())

        # ── 开始转换 ──
        run_btn = tk.Button(right,
                            text="开始转换  →",
                            command=self._run,
                            bg=DARK,
                            fg="#FFFFFF",
                            font=("微软雅黑", 11, "bold"),
                            relief="flat",
                            cursor="hand2",
                            activebackground="#444444",
                            activeforeground="#FFFFFF",
                            pady=14)
        run_btn.pack(fill="x", pady=(14, 0))
        run_btn.bind("<Enter>", lambda e: run_btn.config(bg="#444444"))
        run_btn.bind("<Leave>", lambda e: run_btn.config(bg=DARK))

        # ── 日志 ──
        tk.Label(right, text="运行日志", bg=BG, fg=MID,
                 font=("微软雅黑", 9)).pack(anchor="w", pady=(16, 4))
        log_card = self._card(right)
        log_card.pack(fill="both", expand=True)
        sb = tk.Scrollbar(log_card)
        sb.pack(side="right", fill="y")
        self._log = tk.Text(log_card,
                            yscrollcommand=sb.set,
                            bg=PANEL,
                            fg=MID,
                            font=("微软雅黑", 8),
                            relief="flat",
                            borderwidth=0,
                            highlightthickness=0,
                            state="disabled",
                            wrap="word")
        self._log.pack(fill="both", expand=True, padx=8, pady=8)
        sb.config(command=self._log.yview)

    def _build_statusbar(self):
        bar = tk.Frame(self, bg="#E8E8E8", height=28)
        bar.pack(fill="x", side="bottom")
        bar.pack_propagate(False)
        self._status_var = tk.StringVar(value="就绪")
        tk.Label(bar,
                 textvariable=self._status_var,
                 bg="#E8E8E8",
                 fg=MID,
                 font=("微软雅黑", 8)).pack(side="left", padx=14, pady=6)

    # ── 控件工厂 ─────────────────────────────

    def _card(self, parent):
        return tk.Frame(parent,
                        bg=PANEL,
                        relief="flat",
                        highlightbackground=BORDER,
                        highlightthickness=1)

    def _btn(self, parent, text, cmd, ghost=False):
        kw = dict(font=("微软雅黑", 9),
                  relief="flat",
                  cursor="hand2",
                  padx=10,
                  pady=4)
        if ghost:
            return tk.Button(parent,
                             text=text,
                             command=cmd,
                             bg=GHOST,
                             fg="#444444",
                             activebackground=BORDER,
                             activeforeground=DARK,
                             **kw)
        return tk.Button(parent,
                         text=text,
                         command=cmd,
                         bg=DARK,
                         fg="#FFFFFF",
                         activebackground="#444444",
                         activeforeground="#FFFFFF",
                         **kw)

    # ── 配置变动回调 ─────────────────────────

    def _on_dir_changed(self, *_):
        self._update_preview()
        save_config(
            {"out_dir": self._dir_var.get().strip() or DEFAULT_OUT_DIR})

    def _on_auto_open_changed(self):
        save_config({"auto_open": self._auto_open_var.get()})

    # ── 文本变动（防抖 600ms）────────────────

    def _on_text_modified(self, event=None):
        self._text_input.edit_modified(False)
        if self._debounce_id:
            self.after_cancel(self._debounce_id)
        self._debounce_id = self.after(600, self._auto_detect)

    def _auto_detect(self):
        text = self._text_input.get("1.0", "end").strip()
        if not text:
            self._detect_label.config(text="")
            self._set_name_auto("交易记录导出")
            return
        rtype, records, _ = detect_and_parse(text)
        if rtype == 0:
            self._detect_label.config(text="⚠ 未识别到有效记录", fg="#CC6600")
            return
        label_map = {1: "✓ 成交记录", 2: "✓ 资金流水"}
        self._detect_label.config(
            text=f"{label_map.get(rtype, '✓ 已识别')}  共 {len(records)} 行",
            fg="#2A7A2A")
        self._set_name_auto(infer_filename(text))

    def _set_name_auto(self, name):
        """程序自动设置文件名，不触发配置保存。"""
        self._suppress_name_save = True
        self._name_var.set(name)
        self._suppress_name_save = False

    def _update_preview(self):
        d = self._dir_var.get().strip() or DEFAULT_OUT_DIR
        n = self._name_var.get().strip() or "交易记录导出"
        self._preview_label.config(text=os.path.join(d, n + ".xlsx"))

    # ── 事件 ─────────────────────────────────

    def _clear_text(self):
        self._text_input.delete("1.0", "end")
        self._detect_label.config(text="")
        self._set_name_auto("交易记录导出")

    def _load_file(self):
        init_dir = self._cfg.get("last_open_dir") or DEFAULT_OPEN_DIR
        paths = filedialog.askopenfilenames(title="选择原始记录文件",
                                            initialdir=init_dir,
                                            filetypes=[
                                                ("支持的文件",
                                                 "*.txt *.csv *.xls *.xlsx"),
                                                ("所有文件", "*.*")
                                            ])
        if not paths:
            return

        # 保存最后一次打开目录
        last_dir = os.path.dirname(paths[-1])
        self._cfg["last_open_dir"] = last_dir
        save_config({"last_open_dir": last_dir})

        all_text = []
        for path in paths:
            ext = os.path.splitext(path)[1].lower()
            try:
                if ext in (".xls", ".xlsx"):
                    df = pd.read_excel(path, dtype=str, header=None).fillna("")
                    all_text.append("\n".join("\t".join(str(v) for v in row)
                                              for _, row in df.iterrows()))
                else:
                    with open(path, "rb") as f:
                        raw = f.read()
                    enc = detect_encoding(raw)
                    all_text.append(raw.decode(enc, errors="replace"))
            except Exception as e:
                messagebox.showerror("读取失败", f"{os.path.basename(path)}\n{e}")
                return

        self._text_input.delete("1.0", "end")
        self._text_input.insert("1.0", "\n\n".join(all_text))
        self._status_var.set(f"已载入 {len(paths)} 个文件")

    def _browse_dir(self):
        init_dir = self._dir_var.get().strip() or DEFAULT_OUT_DIR
        d = filedialog.askdirectory(title="选择输出目录", initialdir=init_dir)
        if d:
            self._dir_var.set(d)
            # _on_dir_changed 由 trace 自动触发，此处无需重复保存

    def _get_out_path(self):
        d = self._dir_var.get().strip() or DEFAULT_OUT_DIR
        n = self._name_var.get().strip() or "交易记录导出"
        return os.path.join(d, n + ".xlsx")

    def _log_write(self, msg):
        self._log.config(state="normal")
        self._log.insert("end", msg + "\n")
        self._log.see("end")
        self._log.config(state="disabled")

    # ── 路径确认（两道关卡，递归）────────────

    @staticmethod
    def _is_file_open(path):
        try:
            with open(path, "a+b"):
                return False
        except IOError:
            return True

    def _resolve_output_path(self, path):
        # 第一关：文件已存在
        if os.path.exists(path):
            choice = messagebox.askyesnocancel(
                "文件已存在", f"「{os.path.basename(path)}」已存在。\n\n是=覆盖　否=另存为　取消=终止")
            if choice is None:
                return None
            if choice is False:
                new_path = self._pick_saveas_path(path)
                return self._resolve_output_path(
                    new_path) if new_path else None

        # 第二关：文件被占用（循环）
        while self._is_file_open(path):
            retry = messagebox.askyesnocancel(
                "文件被占用", f"「{os.path.basename(path)}」正被其他程序打开，无法写入。\n\n"
                "请关闭该文件后点击「是」重试\n否=另存为　取消=终止")
            if retry is None:
                return None
            if retry is False:
                new_path = self._pick_saveas_path(path)
                return self._resolve_output_path(
                    new_path) if new_path else None

        return path

    def _pick_saveas_path(self, current_path):
        path = filedialog.asksaveasfilename(
            initialdir=os.path.dirname(current_path),
            initialfile=os.path.basename(current_path),
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")])
        return path if path else None

    # ── 主流程 ───────────────────────────────

    def _run(self):
        text = self._text_input.get("1.0", "end").strip()
        if not text:
            messagebox.showwarning("提示", "请先输入或载入原始记录内容")
            return

        out_path = self._resolve_output_path(self._get_out_path())
        if out_path is None:
            self._status_var.set("已取消")
            return

        self._set_name_auto(os.path.splitext(os.path.basename(out_path))[0])

        self._status_var.set("处理中…")
        self.update_idletasks()

        ok, msg = process_text(text, out_path)
        self._log_write(msg)
        self._status_var.set("✅ 完成" if ok else "❌ 出错")
        (messagebox.showinfo if ok else messagebox.showerror)(
            "完成" if ok else "错误", msg)

        if ok and self._auto_open_var.get():
            os.startfile(out_path)


# ─────────────────────────────────────────────
#  入口
# ─────────────────────────────────────────────

if __name__ == "__main__":
    App().mainloop()
