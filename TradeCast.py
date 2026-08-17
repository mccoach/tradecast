import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
import os
import sys
import re
import json
import shutil
from datetime import datetime
import chardet
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

APP_NAME = "TradeCast v1.4.1-20260817"
CONFIG_FILENAME = "TradeCast_config.json"

DEFAULT_OUT_DIR = r""
DEFAULT_OPEN_DIR = r""

BG = "#F0F0F0"
PANEL = "#FFFFFF"
DARK = "#46515C"
MID = "#666666"
GHOST = "#F7F7F7"
BORDER = "#D8D8D8"
SHADOW = "#C9CED3"
BTN_DARK = "#53606B"
BTN_DARK_ACTIVE = "#46515C"
BTN_GHOST = "#FAFAFA"
BTN_GHOST_ACTIVE = "#EEF1F3"

TYPE1_REQUIRED = {
    "成交日期", "成交时间", "证券代码", "证券名称", "委托类别", "成交价格", "成交数量", "发生金额", "股东代码"
}

TYPE2_REQUIRED = {"成交日期", "成交时间", "委托类别", "发生金额", "摘要", "资金帐号"}

TYPE2_EXCLUDE_PATTERNS = [
    r"[A-Z0-9]{4,}[买卖]",
    r"基金申购",
    r"红利到[帐账]",
    r"确认金额",
    r"配股认购.*确认张数",
    r"领\d{6}红利\d+股\s*\*",
]

OUTPUT_HEADERS = [
    "成交日期", "成交时间", "证券代码", "证券名称", "操作", "成交均价", "成交数量", "股东帐户", "备注"
]
COL_WIDTHS = [12, 10, 14, 24, 16, 12, 10, 14, 52]
LEFT_COLS = {3, 4, 8, 9}

VALID_SOURCE_TYPES = {"TYPE1", "TYPE2"}
VALID_MATCH_TYPES = {"exact", "contains", "regex"}
RULE_FIELDS = {
    "id", "enabled", "source_type", "field", "match_type", "pattern",
    "asset_type", "name_contains", "summary_contains", "output_op",
    "invalid_reason"
}
REQUIRED_RULE_FIELDS = {
    "id", "enabled", "source_type", "field", "match_type", "pattern",
    "output_op"
}

DEFAULT_OPERATION_CATEGORIES = [
    "银证转入",
    "银证转出",
    "股票买入",
    "股票卖出",
    "基金买入",
    "基金卖出",
    "基金申购",
    "基金赎回",
    "理财收益",
    "利息归本",
    "股息红利",
    "红利扣税",
    "转债认购",
    "转债买入",
    "转债卖出",
]

QUANTITY_DIRECTION_ZERO = 0
QUANTITY_DIRECTION_POSITIVE = 1
QUANTITY_DIRECTION_NEGATIVE = -1

QUANTITY_DIRECTION_BY_OPERATION = {
    "银证转入": QUANTITY_DIRECTION_ZERO,
    "银证转出": QUANTITY_DIRECTION_ZERO,
    "理财收益": QUANTITY_DIRECTION_ZERO,
    "利息归本": QUANTITY_DIRECTION_ZERO,
    "股息红利": QUANTITY_DIRECTION_ZERO,
    "红利扣税": QUANTITY_DIRECTION_ZERO,
    "股票买入": QUANTITY_DIRECTION_POSITIVE,
    "基金买入": QUANTITY_DIRECTION_POSITIVE,
    "基金申购": QUANTITY_DIRECTION_POSITIVE,
    "转债认购": QUANTITY_DIRECTION_POSITIVE,
    "转债买入": QUANTITY_DIRECTION_POSITIVE,
    "股票卖出": QUANTITY_DIRECTION_NEGATIVE,
    "基金卖出": QUANTITY_DIRECTION_NEGATIVE,
    "基金赎回": QUANTITY_DIRECTION_NEGATIVE,
    "转债卖出": QUANTITY_DIRECTION_NEGATIVE,
}

DEFAULT_UNKNOWN_POLICY = {
    "unknown_operation_text": "【待手动维护】",
    "unknown_security_text": "【待确认】",
    "unknown_market_prefix": "??",
}

DEFAULT_SETTINGS = {
    "out_dir": DEFAULT_OUT_DIR,
    "last_open_dir": DEFAULT_OPEN_DIR,
    "auto_open": True,
    "window_geometry": "960x700+80+40",
    "window_state": "zoomed",
}

DEFAULT_OPERATION_RULES = [
    {
        "id": "type1_buy_stock",
        "enabled": True,
        "source_type": "TYPE1",
        "field": "委托类别",
        "match_type": "exact",
        "pattern": "买入",
        "asset_type": "股票",
        "output_op": "股票买入"
    },
    {
        "id": "type1_sell_stock",
        "enabled": True,
        "source_type": "TYPE1",
        "field": "委托类别",
        "match_type": "exact",
        "pattern": "卖出",
        "asset_type": "股票",
        "output_op": "股票卖出"
    },
    {
        "id": "type1_buy_fund",
        "enabled": True,
        "source_type": "TYPE1",
        "field": "委托类别",
        "match_type": "exact",
        "pattern": "买入",
        "asset_type": "基金",
        "output_op": "基金买入"
    },
    {
        "id": "type1_sell_fund",
        "enabled": True,
        "source_type": "TYPE1",
        "field": "委托类别",
        "match_type": "exact",
        "pattern": "卖出",
        "asset_type": "基金",
        "output_op": "基金卖出"
    },
    {
        "id": "type1_buy_cbond",
        "enabled": True,
        "source_type": "TYPE1",
        "field": "委托类别",
        "match_type": "exact",
        "pattern": "买入",
        "asset_type": "转债",
        "output_op": "转债买入"
    },
    {
        "id": "type1_sell_cbond",
        "enabled": True,
        "source_type": "TYPE1",
        "field": "委托类别",
        "match_type": "exact",
        "pattern": "卖出",
        "asset_type": "转债",
        "output_op": "转债卖出"
    },
    {
        "id": "type1_fund_subscribe",
        "enabled": True,
        "source_type": "TYPE1",
        "field": "委托类别",
        "match_type": "exact",
        "pattern": "基金申购",
        "output_op": "基金申购"
    },
    {
        "id": "type1_fund_redeem",
        "enabled": True,
        "source_type": "TYPE1",
        "field": "委托类别",
        "match_type": "contains",
        "pattern": "赎回",
        "asset_type": "基金",
        "output_op": "基金赎回"
    },
    {
        "id": "type1_dividend",
        "enabled": True,
        "source_type": "TYPE1",
        "field": "委托类别",
        "match_type": "regex",
        "pattern": "红利|股息|派息",
        "output_op": "股息红利"
    },
    {
        "id": "type1_cbond_subscribe",
        "enabled": True,
        "source_type": "TYPE1",
        "field": "委托类别",
        "match_type": "exact",
        "pattern": "缴款",
        "name_contains": "配债",
        "output_op": "转债认购"
    },
    {
        "id": "type2_bank_in",
        "enabled": True,
        "source_type": "TYPE2",
        "field": "摘要",
        "match_type": "regex",
        "pattern": "存管.*转入|转入.*存管",
        "output_op": "银证转入"
    },
    {
        "id": "type2_bank_out",
        "enabled": True,
        "source_type": "TYPE2",
        "field": "摘要",
        "match_type": "regex",
        "pattern": "存管.*转出|转出.*存管",
        "output_op": "银证转出"
    },
    {
        "id": "type2_interest",
        "enabled": True,
        "source_type": "TYPE2",
        "field": "摘要",
        "match_type": "regex",
        "pattern": "结息|利息",
        "output_op": "利息归本"
    },
    {
        "id": "type2_dividend_tax",
        "enabled": True,
        "source_type": "TYPE2",
        "field": "摘要",
        "match_type": "contains",
        "pattern": "红利差别税扣款",
        "output_op": "红利扣税"
    },
    {
        "id": "type2_wealth_income",
        "enabled": True,
        "source_type": "TYPE2",
        "field": "摘要",
        "match_type": "contains",
        "pattern": "理财收益",
        "output_op": "理财收益"
    },
]

_SECURITY_PREFIX_RULES_RAW = [
    ("689", "SH", "股票", "科创板股票"),
    ("688", "SH", "股票", "科创板股票"),
    ("605", "SH", "股票", "主板A股"),
    ("603", "SH", "股票", "主板A股"),
    ("601", "SH", "股票", "主板A股"),
    ("600", "SH", "股票", "主板A股"),
    ("900", "SH", "股票", "B股"),
    ("360", "SH", "股票", "优先股"),
    ("330", "SH", "股票", "优先股"),
    ("789", "SH", "股票", "主板A股"),
    ("787", "SH", "股票", "主板A股"),
    ("736", "SH", "股票", "主板A股"),
    ("732", "SH", "股票", "主板A股"),
    ("30", "SZ", "股票", "创业板股票"),
    ("004", "SZ", "股票", "主板A股"),
    ("003", "SZ", "股票", "主板A股"),
    ("002", "SZ", "股票", "主板A股"),
    ("001", "SZ", "股票", "主板A股"),
    ("000", "SZ", "股票", "主板A股"),
    ("20", "SZ", "股票", "B股"),
    ("140", "SZ", "股票", "优先股"),
    ("589", "SH", "基金", "场内ETF"),
    ("588", "SH", "基金", "场内ETF"),
    ("563", "SH", "基金", "场内ETF"),
    ("562", "SH", "基金", "场内ETF"),
    ("561", "SH", "基金", "场内ETF"),
    ("560", "SH", "基金", "场内ETF"),
    ("551", "SH", "基金", "场内ETF"),
    ("530", "SH", "基金", "场内ETF"),
    ("526", "SH", "基金", "场内ETF"),
    ("5209", "SH", "基金", "场内ETF"),
    ("5208", "SH", "基金", "场内ETF"),
    ("5207", "SH", "基金", "场内ETF"),
    ("5206", "SH", "基金", "场内ETF"),
    ("5205", "SH", "基金", "场内ETF"),
    ("518", "SH", "基金", "场内ETF"),
    ("517", "SH", "基金", "场内ETF"),
    ("516", "SH", "基金", "场内ETF"),
    ("515", "SH", "基金", "场内ETF"),
    ("513", "SH", "基金", "场内ETF"),
    ("512", "SH", "基金", "场内ETF"),
    ("511", "SH", "基金", "场内ETF"),
    ("510", "SH", "基金", "场内ETF"),
    ("5060", "SH", "基金", "LOF基金"),
    ("502", "SH", "基金", "LOF基金"),
    ("501", "SH", "基金", "LOF基金"),
    ("5080", "SH", "基金", "公募REITs"),
    ("550", "SH", "基金", "封闭式基金"),
    ("5058", "SH", "基金", "封闭式基金"),
    ("500", "SH", "基金", "封闭式基金"),
    ("519", "SH", "基金", "开放式基金"),
    ("159", "SZ", "基金", "场内ETF"),
    ("158", "SZ", "基金", "场内ETF"),
    ("17", "SZ", "基金", "LOF基金"),
    ("16", "SZ", "基金", "LOF基金"),
    ("151", "SZ", "基金", "LOF基金"),
    ("150", "SZ", "基金", "LOF基金"),
    ("1819", "SZ", "基金", "公募REITs"),
    ("1818", "SZ", "基金", "公募REITs"),
    ("1817", "SZ", "基金", "公募REITs"),
    ("1816", "SZ", "基金", "公募REITs"),
    ("1815", "SZ", "基金", "公募REITs"),
    ("1814", "SZ", "基金", "公募REITs"),
    ("1813", "SZ", "基金", "公募REITs"),
    ("1812", "SZ", "基金", "公募REITs"),
    ("1811", "SZ", "基金", "公募REITs"),
    ("1810", "SZ", "基金", "公募REITs"),
    ("1809", "SZ", "基金", "公募REITs"),
    ("1808", "SZ", "基金", "公募REITs"),
    ("1807", "SZ", "基金", "公募REITs"),
    ("1806", "SZ", "基金", "公募REITs"),
    ("1805", "SZ", "基金", "公募REITs"),
    ("1804", "SZ", "基金", "公募REITs"),
    ("1803", "SZ", "基金", "公募REITs"),
    ("1802", "SZ", "基金", "公募REITs"),
    ("1801", "SZ", "基金", "公募REITs"),
    ("184", "SZ", "基金", "封闭式基金"),
]
SECURITY_PREFIX_RULES = sorted(_SECURITY_PREFIX_RULES_RAW,
                               key=lambda x: len(x[0]),
                               reverse=True)


def program_dir():
    return os.path.dirname(
        os.path.abspath(
            sys.executable if getattr(sys, "frozen", False) else __file__))


def _config_path():
    return os.path.join(program_dir(), CONFIG_FILENAME)


def clean_path(value):
    s = str(value or "").strip()
    pairs = [('"', '"'), ("'", "'"), ("\u201c", "\u201d"),
             ("\u2018", "\u2019")]
    changed = True
    while changed and len(s) >= 2:
        changed = False
        for l, r in pairs:
            if s.startswith(l) and s.endswith(r):
                s = s[1:-1].strip()
                changed = True
                break
    return os.path.expandvars(os.path.expanduser(s)).replace("/", "\\")


def default_config():
    return {
        "config_version": 1,
        "settings": dict(DEFAULT_SETTINGS),
        "operation_categories": list(DEFAULT_OPERATION_CATEGORIES),
        "operation_rules": [dict(r) for r in DEFAULT_OPERATION_RULES],
        "unknown_policy": dict(DEFAULT_UNKNOWN_POLICY),
    }


def backup_broken_config(path):
    if os.path.exists(path):
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        try:
            shutil.copy2(
                path,
                os.path.join(os.path.dirname(path),
                             f"TradeCast_config.broken_{stamp}.json"))
        except Exception:
            pass


def unique_id(base, used):
    base = re.sub(r"\W+", "_", str(base).strip()) or "rule"
    candidate = base
    i = 2
    while candidate in used:
        candidate = f"{base}_{i}"
        i += 1
    used.add(candidate)
    return candidate


def normalize_rule(raw, used_ids):
    raw = raw if isinstance(raw, dict) else {}
    rule = {k: raw.get(k) for k in RULE_FIELDS if k in raw}
    if "enabled" not in rule or not isinstance(rule.get("enabled"), bool):
        rule["enabled"] = False
        rule["invalid_reason"] = "enabled字段非法，已禁用"
    rule["id"] = unique_id(rule.get("id") or "rule", used_ids)

    bad = []
    for k in REQUIRED_RULE_FIELDS:
        if k != "enabled" and not str(rule.get(k, "")).strip():
            bad.append(f"缺少{k}")
    if rule.get("source_type") not in VALID_SOURCE_TYPES:
        bad.append("source_type非法")
    if rule.get("match_type") not in VALID_MATCH_TYPES:
        bad.append("match_type非法")

    if bad:
        rule["enabled"] = False
        rule["invalid_reason"] = "；".join(bad)
    else:
        rule.pop("invalid_reason", None)

    for k in ("asset_type", "name_contains", "summary_contains"):
        if not str(rule.get(k, "")).strip():
            rule.pop(k, None)
    return rule


def normalize_config(cfg):
    cfg = cfg if isinstance(cfg, dict) else {}
    clean = default_config()
    src_settings = cfg.get("settings") if isinstance(cfg.get("settings"),
                                                     dict) else {}

    for k, dv in DEFAULT_SETTINGS.items():
        v = src_settings.get(k, dv)
        if isinstance(dv, bool):
            clean["settings"][k] = bool(v)
        elif k in ("out_dir", "last_open_dir"):
            clean["settings"][k] = clean_path(v)
        else:
            clean["settings"][k] = str(v or dv)

    cats = cfg.get("operation_categories")
    cats = cats if isinstance(cats, list) else []
    seen = set()
    clean["operation_categories"] = []
    for c in cats + DEFAULT_OPERATION_CATEGORIES:
        c = str(c).strip()
        if c and c not in seen:
            seen.add(c)
            clean["operation_categories"].append(c)

    used = set()
    clean["operation_rules"] = []
    src_rules = cfg.get("operation_rules") if isinstance(
        cfg.get("operation_rules"), list) else []
    for r in src_rules:
        nr = normalize_rule(r, used)
        clean["operation_rules"].append(nr)
        op = nr.get("output_op")
        if op and op not in clean["operation_categories"]:
            clean["operation_categories"].append(op)

    ids = {r.get("id") for r in clean["operation_rules"]}
    for r in DEFAULT_OPERATION_RULES:
        if r["id"] not in ids:
            clean["operation_rules"].append(normalize_rule(dict(r), used))

    src_policy = cfg.get("unknown_policy") if isinstance(
        cfg.get("unknown_policy"), dict) else {}
    clean["unknown_policy"] = {
        k: str(src_policy.get(k, v) or v)
        for k, v in DEFAULT_UNKNOWN_POLICY.items()
    }
    clean["config_version"] = 1
    return clean


def save_config(cfg):
    cfg = normalize_config(cfg)
    with open(_config_path(), "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)
    return cfg


def load_config():
    path = _config_path()
    if not os.path.exists(path):
        return save_config(default_config())
    try:
        with open(path, "r", encoding="utf-8") as f:
            cfg = json.load(f)
    except Exception:
        backup_broken_config(path)
        return save_config(default_config())
    return save_config(normalize_config(cfg))


def detect_encoding(raw_bytes):
    enc = chardet.detect(raw_bytes).get("encoding") or "utf-8"
    return "gb18030" if enc.lower() in ("gb2312", "gbk", "gb18030") else enc


def to_float(val, default=0.0):
    try:
        return float(str(val).replace(",", "").strip())
    except Exception:
        return default


def to_int_qty(val):
    try:
        return int(float(str(val).replace(",", "").strip()))
    except Exception:
        return 0


def resolve_output_quantity(op_key, raw_qty):
    qty_abs = abs(to_int_qty(raw_qty))
    direction = QUANTITY_DIRECTION_BY_OPERATION.get(op_key)

    if direction == QUANTITY_DIRECTION_ZERO:
        return 0, ""

    if direction == QUANTITY_DIRECTION_POSITIVE:
        return qty_abs, ""

    if direction == QUANTITY_DIRECTION_NEGATIVE:
        return -qty_abs, ""

    return 0, f"操作类别未配置数量方向：{op_key}，成交数量已输出为0"


def fmt_date(val):
    s = str(val).strip().replace("-", "").replace("/", "")
    return f"{s[:4]}-{s[4:6]}-{s[6:]}" if len(s) == 8 and s.isdigit() else s


def normalize_code(code):
    s = str(code).strip()
    if not s:
        return ""
    if re.fullmatch(r"\d+(\.0+)?", s):
        s = str(int(float(s)))
    return s.zfill(6)


def infer_market_by_account(account):
    s = str(account).strip()
    if not s:
        return None
    if re.match(r"^[A-Za-z]", s):
        return "SH"
    if re.match(r"^\d", s):
        return "SZ"
    return None


def lookup_prefix_security(code):
    code = normalize_code(code)
    for prefix, market, asset_type, desc in SECURITY_PREFIX_RULES:
        if code.startswith(prefix):
            return {
                "market": market,
                "asset_type": asset_type,
                "desc": desc,
                "prefix": prefix
            }
    return None


def resolve_security(code, name, account, unknown_policy):
    code6 = normalize_code(code)
    name = str(name).strip()
    account_market = infer_market_by_account(account)
    prefix_hit = lookup_prefix_security(code6)
    remarks = []

    if prefix_hit:
        market = prefix_hit["market"]
        asset_type = prefix_hit["asset_type"]
        if account_market and account_market != market:
            remarks.append(
                f"市场校核冲突：代码前缀识别为{market}，股东账户推断为{account_market}，已按代码前缀输出")
    else:
        market = account_market or unknown_policy["unknown_market_prefix"]
        asset_type = "转债" if name and ("转债" in name or "配债" in name) else "未知"
        if not account_market:
            remarks.append("证券市场无法识别，代码已使用??前缀")

    return {
        "market": market,
        "asset_type": asset_type,
        "code": f"{market}{code6}" if code6 else "",
        "remarks": remarks,
        "prefix_hit": bool(prefix_hit),
    }


def join_remarks(items):
    return "；".join([str(x).strip() for x in items if str(x).strip()])


def extract_code_from_dividend_tax_summary(summary):
    m = re.search(r"证券代码[:：]\s*(\d{6})", str(summary))
    return normalize_code(m.group(1)) if m else ""


def parse_text(text):
    lines = [l for l in text.splitlines() if l.strip()]
    rows = []
    for line in lines:
        parts = line.split("\t") if "\t" in line else re.split(r" {2,}", line)
        rows.append([p.strip() for p in parts if p.strip()])
    return rows


def detect_header(row):
    row_set = set(row)
    if TYPE1_REQUIRED.issubset(row_set):
        return {
            "rtype": 1,
            "col_map": {
                field: row.index(field)
                for field in TYPE1_REQUIRED
            }
        }
    if TYPE2_REQUIRED.issubset(row_set):
        return {
            "rtype": 2,
            "col_map": {
                field: row.index(field)
                for field in TYPE2_REQUIRED
            }
        }
    return None


def build_records_from_block(rows, start_idx, end_idx, col_map):
    records = []
    for row in rows[start_idx + 1:end_idx]:
        if not any(row) or detect_header(row):
            continue
        record = {
            field: (row[idx] if idx < len(row) else "")
            for field, idx in col_map.items()
        }
        if any(str(v).strip() for v in record.values()):
            records.append(record)
    return records


def parse_record_blocks(text):
    rows = parse_text(text)
    headers = []
    for idx, row in enumerate(rows):
        header = detect_header(row)
        if header:
            headers.append({
                "idx": idx,
                "rtype": header["rtype"],
                "col_map": header["col_map"]
            })

    blocks = []
    for i, header in enumerate(headers):
        end_idx = headers[i + 1]["idx"] if i + 1 < len(headers) else len(rows)
        records = build_records_from_block(rows, header["idx"], end_idx,
                                           header["col_map"])
        if records:
            blocks.append({
                "rtype": header["rtype"],
                "records": records,
                "col_map": header["col_map"]
            })
    return blocks


def infer_filename(text):
    blocks = parse_record_blocks(text)
    if not blocks:
        return "交易记录导出"

    dates = []
    has_type1 = False
    has_type2 = False
    for block in blocks:
        has_type1 = has_type1 or block["rtype"] == 1
        has_type2 = has_type2 or block["rtype"] == 2
        for r in block["records"]:
            d = str(r.get("成交日期", "")).strip().replace("-",
                                                       "").replace("/", "")
            if len(d) == 8 and d.isdigit():
                dates.append(d)

    if dates:
        d_min, d_max = min(dates), max(dates)
        date_range = f"{d_min}-{d_max}" if d_min != d_max else d_min
    else:
        date_range = "未知日期"

    label = "交易及资金记录" if has_type1 and has_type2 else (
        "转账记录" if has_type2 else "交易记录")
    return f"{date_range}{label}"


def field_match(value, match_type, pattern):
    value = str(value or "")
    pattern = str(pattern or "")
    if match_type == "exact":
        return value == pattern
    if match_type == "contains":
        return pattern in value
    if match_type == "regex":
        try:
            return re.search(pattern, value) is not None
        except re.error:
            return False
    return False


def match_operation(ctx, rules):
    for rule in rules:
        if not rule.get("enabled") or rule.get("invalid_reason") or rule.get(
                "source_type") != ctx.get("source_type"):
            continue
        if not field_match(ctx.get(rule.get("field", ""), ""),
                           rule.get("match_type"), rule.get("pattern")):
            continue
        asset_type = str(rule.get("asset_type", "")).strip()
        if asset_type and asset_type != str(ctx.get("asset_type", "")).strip():
            continue
        name_contains = str(rule.get("name_contains", "")).strip()
        if name_contains and name_contains not in str(ctx.get("证券名称", "")):
            continue
        summary_contains = str(rule.get("summary_contains", "")).strip()
        if summary_contains and summary_contains not in str(ctx.get("摘要", "")):
            continue
        return rule.get("output_op")
    return None


def process_type1(records, config):
    rules = config["operation_rules"]
    policy = config["unknown_policy"]
    results = []
    warnings = []
    zero_filtered = 0

    for row in records:
        amt = to_float(row.get("发生金额", 0))
        if amt == 0:
            zero_filtered += 1
            continue

        remarks = []
        op_raw = str(row.get("委托类别", "")).strip()
        code_raw = normalize_code(row.get("证券代码", ""))
        name = str(row.get("证券名称",
                           "")).strip() or policy["unknown_security_text"]
        date = fmt_date(row.get("成交日期", ""))
        time_ = str(row.get("成交时间", "")).strip()
        price_raw = str(row.get("成交价格", "")).strip()
        qty_raw = row.get("成交数量", "0")
        account = str(row.get("股东代码", "")).strip()
        deal_no = str(row.get("成交编号", "")).strip()

        sec = resolve_security(code_raw, name, account, policy)
        remarks.extend(sec["remarks"])
        asset_type = sec["asset_type"]

        ctx = {
            "source_type": "TYPE1",
            "委托类别": op_raw,
            "证券代码": code_raw,
            "证券名称": name,
            "成交编号": deal_no,
            "asset_type": asset_type,
        }
        op_key = match_operation(ctx, rules)

        if op_key == "转债认购":
            remarks.append("配债认购按原始配债代码输出，可能需手动维护为正式转债代码")

        if not op_key:
            op_key = policy["unknown_operation_text"]
            remarks.append(f"操作无法识别，已标记为{op_key}")
            if asset_type == "未知":
                remarks.append("标的类型无法识别，操作需手动维护")
            warnings.append(
                f"{date} {code_raw} {name} 委托类别={op_raw} 发生金额={amt}")

        out_qty, qty_remark = resolve_output_quantity(op_key, qty_raw)
        if qty_remark:
            remarks.append(qty_remark)

        if out_qty == 0:
            price = abs(amt)
        else:
            price = to_float(price_raw, abs(amt))
            if price == 0 and amt != 0:
                price = abs(amt)

        results.append({
            "成交日期": date,
            "成交时间": time_,
            "证券代码": sec["code"],
            "证券名称": name,
            "操作_公式": op_key,
            "成交均价": price,
            "成交数量": out_qty,
            "股东帐户": account,
            "备注": join_remarks(remarks),
        })

    return results, {"unknown_ops": warnings, "zero_filtered": zero_filtered}


def process_type2(records, config):
    rules = config["operation_rules"]
    policy = config["unknown_policy"]
    results = []
    unknowns = []
    excluded_summaries = []
    zero_filtered = 0

    for row in records:
        amount = to_float(row.get("发生金额", 0))
        if amount == 0:
            zero_filtered += 1
            continue

        summary = str(row.get("摘要", "")).strip()
        if any(re.search(pat, summary) for pat in TYPE2_EXCLUDE_PATTERNS):
            excluded_summaries.append(summary)
            continue

        date = fmt_date(row.get("成交日期", ""))
        time_ = str(row.get("成交时间", "")).strip()

        ctx = {
            "source_type": "TYPE2",
            "摘要": summary,
            "委托类别": str(row.get("委托类别", "")).strip(),
            "asset_type": "",
        }
        op_key = match_operation(ctx, rules)
        remarks = []

        if not op_key:
            op_key = policy["unknown_operation_text"]
            remarks.append(f"摘要无法匹配操作规则，已标记为{op_key}")
            unknowns.append(f"{date} {summary} 发生金额={amount}")

        code_val = ""
        name_val = ""

        if op_key in ("银证转入", "银证转出"):
            name_val = "XLOOKUP(INDEX(XLOOKUP(\"资金账户\",$1:$1,$1:$10000),ROW()),XLOOKUP(\"所属资金账户\",$1:$1,$1:$100),XLOOKUP(\"所属存管账户\",$1:$1,$1:$100))"
        elif op_key == "利息归本":
            name_val = ""
        elif op_key == "红利扣税":
            remarks.append("红利差别税扣款，按红利扣税输出")
            code_raw = extract_code_from_dividend_tax_summary(summary)
            if code_raw:
                sec = resolve_security(code_raw, "", "", policy)
                code_val = sec["code"]
                name_val = policy["unknown_security_text"]
                remarks.extend(sec["remarks"])
                remarks.append("已从摘要提取证券代码，证券名称需确认")
            else:
                remarks.append("未能从摘要提取证券代码")

        results.append({
            "成交日期": date,
            "成交时间": time_,
            "证券代码": code_val,
            "证券名称": name_val,
            "操作_公式": op_key,
            "成交均价": abs(amount),
            "成交数量": 0,
            "股东帐户": "",
            "备注": join_remarks(remarks),
        })

    return results, {
        "unknown_summaries": unknowns,
        "excluded_summaries": excluded_summaries,
        "zero_filtered": zero_filtered,
    }


def write_xlsx(all_rows, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "交易记录"

    header_fill = PatternFill("solid", fgColor="2B2B2B")
    header_font = Font(color="FFFFFF", bold=True, name="微软雅黑", size=10)
    even_fill = PatternFill("solid", fgColor="F5F5F5")
    odd_fill = PatternFill("solid", fgColor="FFFFFF")
    border_side = Side(style="thin", color="DDDDDD")
    cell_border = Border(left=border_side,
                         right=border_side,
                         top=border_side,
                         bottom=border_side)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    remark_align = Alignment(horizontal="left",
                             vertical="center",
                             wrap_text=True)

    for col, h in enumerate(OUTPUT_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center_align
        c.border = cell_border

    for row_idx, r in enumerate(all_rows, 2):
        fill = even_fill if row_idx % 2 == 0 else odd_fill

        def wc(col, value, num_fmt=None):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill = fill
            cell.border = cell_border
            cell.alignment = remark_align if col == 9 else (
                left_align if col in LEFT_COLS else center_align)
            if num_fmt:
                cell.number_format = num_fmt

        wc(1, r["成交日期"], "YYYY-MM-DD")
        wc(2, r["成交时间"], "HH:MM:SS")
        wc(3, r["证券代码"])
        wc(4, r["证券名称"])
        wc(5, r["操作_公式"])
        wc(6, r["成交均价"], "0.0000")
        wc(7, r["成交数量"], "0")
        wc(8, r["股东帐户"], "@")
        wc(9, r.get("备注", ""))

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(out_path)


def process_text(text, out_path, config):
    blocks = parse_record_blocks(text)
    if not blocks:
        return False, "无法识别记录类型，请检查内容格式"

    all_rows = []
    info_total = {
        "unknown_ops": [],
        "unknown_summaries": [],
        "excluded_summaries": [],
        "zero_filtered": 0,
        "type1_blocks": 0,
        "type2_blocks": 0,
        "type1_records": 0,
        "type2_records": 0,
    }

    for block in blocks:
        records = block["records"]
        if block["rtype"] == 1:
            rows, info = process_type1(records, config)
            info_total["type1_blocks"] += 1
            info_total["type1_records"] += len(records)
        elif block["rtype"] == 2:
            rows, info = process_type2(records, config)
            info_total["type2_blocks"] += 1
            info_total["type2_records"] += len(records)
        else:
            continue

        all_rows.extend(rows)
        info_total["zero_filtered"] += info.get("zero_filtered", 0)
        info_total["unknown_ops"].extend(info.get("unknown_ops", []))
        info_total["unknown_summaries"].extend(
            info.get("unknown_summaries", []))
        info_total["excluded_summaries"].extend(
            info.get("excluded_summaries", []))

    write_xlsx(all_rows, out_path)

    warnings = [
        f"已识别成交记录区块：{info_total['type1_blocks']} 个，原始记录 {info_total['type1_records']} 条",
        f"已识别资金流水区块：{info_total['type2_blocks']} 个，原始记录 {info_total['type2_records']} 条",
    ]

    if info_total["zero_filtered"]:
        warnings.append(f"已过滤发生金额为 0 的记录：{info_total['zero_filtered']} 条")
    if info_total["unknown_ops"]:
        warnings.append("⚠️ 以下成交记录操作无法识别，已在操作列标记为【待手动维护】：\n" +
                        "\n".join(info_total["unknown_ops"][:20]))
    if info_total["unknown_summaries"]:
        warnings.append("⚠️ 以下资金流水摘要无法识别，已在操作列标记为【待手动维护】：\n" +
                        "\n".join(info_total["unknown_summaries"][:20]))
    if info_total["excluded_summaries"]:
        uniq = list(dict.fromkeys(info_total["excluded_summaries"]))[:20]
        warnings.append("🔕 以下摘要已主动过滤：\n" + "\n".join(uniq))

    return True, f"✅ 导出成功！共 {len(all_rows)} 条\n路径：{out_path}\n\n" + "\n\n".join(
        warnings)


class RuleEditor(tk.Toplevel):
    SOURCE_TYPE_TO_LABEL = {"TYPE1": "成交记录", "TYPE2": "资金流水"}
    LABEL_TO_SOURCE_TYPE = {v: k for k, v in SOURCE_TYPE_TO_LABEL.items()}
    MATCH_TYPE_TO_LABEL = {"exact": "等于", "contains": "包含", "regex": "正则"}
    LABEL_TO_MATCH_TYPE = {v: k for k, v in MATCH_TYPE_TO_LABEL.items()}

    def __init__(self, master, cfg, rule=None, output_op=None):
        super().__init__(master)
        self.title("编辑规则")
        self.geometry("560x500")
        self.resizable(False, False)
        self.cfg = cfg
        self.result = None
        self.rule = dict(rule) if rule else {
            "id": "",
            "enabled": True,
            "source_type": "TYPE1",
            "field": "委托类别",
            "match_type": "exact",
            "pattern": "",
            "asset_type": "",
            "name_contains": "",
            "summary_contains": "",
            "output_op": output_op or "",
        }
        self.vars = {}
        self._build()
        self.bind("<Escape>", lambda e: self.destroy())
        self.grab_set()

    def _label_value(self, key, value):
        return self.SOURCE_TYPE_TO_LABEL.get(
            value,
            value) if key == "source_type" else self.MATCH_TYPE_TO_LABEL.get(
                value, value) if key == "match_type" else value

    def _internal_value(self, key, value):
        return self.LABEL_TO_SOURCE_TYPE.get(
            value,
            value) if key == "source_type" else self.LABEL_TO_MATCH_TYPE.get(
                value, value) if key == "match_type" else value

    def _row(self, parent, label, key, values=None):
        row = tk.Frame(parent)
        row.pack(fill="x", padx=14, pady=5)
        tk.Label(row, text=label, width=18, anchor="e").pack(side="left")
        var = tk.StringVar(
            value=self._label_value(key, str(self.rule.get(key, "") or "")))
        self.vars[key] = var
        if values:
            ttk.Combobox(row,
                         textvariable=var,
                         values=values,
                         state="readonly").pack(side="left",
                                                fill="x",
                                                expand=True)
        else:
            tk.Entry(row, textvariable=var).pack(side="left",
                                                 fill="x",
                                                 expand=True)

    def _build(self):
        body = tk.Frame(self)
        body.pack(fill="both", expand=True, pady=10)
        tk.Label(body, text="当记录满足主要匹配条件和附加条件时，输出为指定操作类别。",
                 fg="#666666").pack(anchor="w", padx=20, pady=(0, 8))
        self.enabled_var = tk.BooleanVar(
            value=bool(self.rule.get("enabled", True)))
        tk.Checkbutton(body, text="是否启用",
                       variable=self.enabled_var).pack(anchor="w", padx=32)
        self._row(body, "规则编号", "id")
        self._row(body, "记录类型", "source_type", ["成交记录", "资金流水"])
        self._row(body, "主要匹配字段", "field", ["委托类别", "摘要", "证券名称", "成交编号"])
        self._row(body, "主要匹配方式", "match_type", ["等于", "包含", "正则"])
        self._row(body, "主要匹配内容", "pattern")
        self._row(body, "附加条件：标的类型", "asset_type",
                  ["", "股票", "基金", "转债", "未知"])
        self._row(body, "附加条件：证券名称包含", "name_contains")
        self._row(body, "附加条件：摘要包含", "summary_contains")
        self._row(body, "输出操作类别", "output_op",
                  self.cfg["operation_categories"])
        btns = tk.Frame(self)
        btns.pack(fill="x", pady=10)
        tk.Button(btns, text="保存", command=self._save, bg=DARK,
                  fg="white").pack(side="right", padx=10)
        tk.Button(btns, text="取消", command=self.destroy).pack(side="right")

    def _save(self):
        rule = {
            k: self._internal_value(k,
                                    v.get().strip())
            for k, v in self.vars.items()
        }
        rule["enabled"] = self.enabled_var.get()

        if not rule["id"]:
            messagebox.showwarning("提示", "规则编号不能为空", parent=self)
            return
        if not rule["field"] or not rule["pattern"] or not rule["output_op"]:
            messagebox.showwarning("提示",
                                   "主要匹配字段、主要匹配内容、输出操作类别不能为空",
                                   parent=self)
            return
        if rule["source_type"] not in VALID_SOURCE_TYPES or rule[
                "match_type"] not in VALID_MATCH_TYPES:
            messagebox.showwarning("提示", "记录类型或匹配方式非法", parent=self)
            return

        for k in ("asset_type", "name_contains", "summary_contains"):
            if not rule.get(k):
                rule.pop(k, None)

        self.result = rule
        self.destroy()


class RuleManager(tk.Toplevel):
    SOURCE_TYPE_LABELS = {"TYPE1": "成交记录", "TYPE2": "资金流水"}
    MATCH_TYPE_LABELS = {"exact": "等于", "contains": "包含", "regex": "正则"}

    def __init__(self, master, cfg, save_callback):
        super().__init__(master)
        self.title("规则维护")
        self.geometry("1040x600")
        self.cfg = cfg
        self.save_callback = save_callback
        self.current_category = None
        self._build()
        self._refresh_categories()
        self.bind("<Escape>", lambda e: self.destroy())
        self.grab_set()

    def _build(self):
        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=12, pady=12)

        left = tk.Frame(body, bg=BG, width=260)
        left.pack(side="left", fill="y")
        left.pack_propagate(False)

        tk.Label(left, text="操作类别", bg=BG,
                 font=("微软雅黑", 10, "bold")).pack(anchor="w")
        self.cat_list = tk.Listbox(left, exportselection=False)
        self.cat_list.pack(fill="both", expand=True, pady=6)
        self.cat_list.bind("<<ListboxSelect>>", self._on_category_select)

        cat_btns = tk.Frame(left, bg=BG)
        cat_btns.pack(fill="x")
        tk.Button(cat_btns, text="新增",
                  command=self._add_category).pack(side="left", padx=2)
        tk.Button(cat_btns, text="重命名",
                  command=self._rename_category).pack(side="left", padx=2)
        tk.Button(cat_btns, text="删除",
                  command=self._delete_category).pack(side="left", padx=2)

        right = tk.Frame(body, bg=BG)
        right.pack(side="right", fill="both", expand=True, padx=(14, 0))
        tk.Label(right, text="映射规则", bg=BG,
                 font=("微软雅黑", 10, "bold")).pack(anchor="w")

        columns = ("enabled", "source_type", "field", "match_type", "pattern",
                   "asset_type", "name_contains", "summary_contains",
                   "invalid_reason")
        self.rule_table = ttk.Treeview(right,
                                       columns=columns,
                                       show="headings",
                                       selectmode="browse")
        headings = {
            "enabled": "启用",
            "source_type": "记录类型",
            "field": "匹配字段",
            "match_type": "匹配方式",
            "pattern": "匹配内容",
            "asset_type": "标的类型",
            "name_contains": "名称包含",
            "summary_contains": "摘要包含",
            "invalid_reason": "状态说明"
        }
        for col in columns:
            self.rule_table.heading(col, text=headings[col])
            self.rule_table.column(col, width=110, anchor="w")
        self.rule_table.pack(fill="both", expand=True, pady=6)

        rule_btns = tk.Frame(right, bg=BG)
        rule_btns.pack(fill="x")
        tk.Button(rule_btns, text="新增规则",
                  command=self._add_rule).pack(side="left", padx=2)
        tk.Button(rule_btns, text="编辑规则",
                  command=self._edit_rule).pack(side="left", padx=2)
        tk.Button(rule_btns, text="删除规则",
                  command=self._delete_rule).pack(side="left", padx=2)
        tk.Button(rule_btns, text="启用/禁用",
                  command=self._toggle_rule).pack(side="left", padx=2)
        tk.Button(rule_btns, text="关闭",
                  command=self.destroy).pack(side="right")

    def _refresh_categories(self):
        self.cat_list.delete(0, "end")
        for c in self.cfg["operation_categories"]:
            self.cat_list.insert("end", c)
        if self.current_category not in self.cfg["operation_categories"]:
            self.current_category = self.cfg["operation_categories"][
                0] if self.cfg["operation_categories"] else None
        if self.current_category:
            idx = self.cfg["operation_categories"].index(self.current_category)
            self.cat_list.selection_set(idx)
        self._refresh_rules()

    def _on_category_select(self, event=None):
        sel = self.cat_list.curselection()
        if sel:
            self.current_category = self.cat_list.get(sel[0])
            self._refresh_rules()

    def _selected_rule_index(self):
        selected = self.rule_table.selection()
        if not selected:
            return None
        try:
            return int(self.rule_table.item(selected[0], "text"))
        except Exception:
            return None

    def _rule_values(self, r):
        return (
            "是" if r.get("enabled") and not r.get("invalid_reason") else "否",
            self.SOURCE_TYPE_LABELS.get(r.get("source_type"),
                                        r.get("source_type", "")),
            r.get("field", ""),
            self.MATCH_TYPE_LABELS.get(r.get("match_type"),
                                       r.get("match_type", "")),
            r.get("pattern", ""),
            r.get("asset_type", "") or "任意",
            r.get("name_contains", ""),
            r.get("summary_contains", ""),
            r.get("invalid_reason", ""),
        )

    def _refresh_rules(self):
        for item in self.rule_table.get_children():
            self.rule_table.delete(item)
        if not self.current_category:
            return
        for idx, r in enumerate(self.cfg["operation_rules"]):
            if r.get("output_op") == self.current_category:
                self.rule_table.insert("",
                                       "end",
                                       text=str(idx),
                                       values=self._rule_values(r))

    def _confirm_fee(self, action):
        return messagebox.askyesno(
            "确认操作类别修改",
            f"{action}操作类别会影响后续费率计算。\n\n请确认已经或将同步修改费率信息。\n\n是否继续？",
            parent=self)

    def _add_category(self):
        if not self._confirm_fee("新增"):
            return
        name = simpledialog.askstring("新增类别", "请输入新操作类别：", parent=self)
        if not name:
            return
        name = name.strip()
        if not name or name in self.cfg["operation_categories"]:
            return
        self.cfg["operation_categories"].append(name)
        self.current_category = name
        self.save_callback()
        self._refresh_categories()

    def _rename_category(self):
        old = self.current_category
        if not old or not self._confirm_fee("重命名"):
            return
        new = simpledialog.askstring("重命名类别",
                                     "请输入新名称：",
                                     initialvalue=old,
                                     parent=self)
        if not new:
            return
        new = new.strip()
        if not new or new == old or new in self.cfg["operation_categories"]:
            return
        self.cfg["operation_categories"] = [
            new if c == old else c for c in self.cfg["operation_categories"]
        ]
        for r in self.cfg["operation_rules"]:
            if r.get("output_op") == old:
                r["output_op"] = new
        self.current_category = new
        self.save_callback()
        self._refresh_categories()

    def _delete_category(self):
        cat = self.current_category
        if not cat or not self._confirm_fee("删除"):
            return
        self.cfg["operation_categories"] = [
            c for c in self.cfg["operation_categories"] if c != cat
        ]
        for r in self.cfg["operation_rules"]:
            if r.get("output_op") == cat:
                r["enabled"] = False
                r["invalid_reason"] = "输出操作类别已删除"
        self.current_category = self.cfg["operation_categories"][
            0] if self.cfg["operation_categories"] else None
        self.save_callback()
        self._refresh_categories()

    def _add_rule(self):
        if not self.current_category:
            return
        dlg = RuleEditor(self, self.cfg, output_op=self.current_category)
        self.wait_window(dlg)
        if not dlg.result:
            return
        if any(
                r.get("id") == dlg.result["id"]
                for r in self.cfg["operation_rules"]):
            messagebox.showwarning("提示", "规则编号已存在", parent=self)
            return
        self.cfg["operation_rules"].append(dlg.result)
        self.save_callback()
        self._refresh_rules()

    def _edit_rule(self):
        idx = self._selected_rule_index()
        if idx is None:
            return
        old_id = self.cfg["operation_rules"][idx].get("id")
        dlg = RuleEditor(self, self.cfg, self.cfg["operation_rules"][idx])
        self.wait_window(dlg)
        if not dlg.result:
            return
        new_id = dlg.result["id"]
        if new_id != old_id and any(
                r.get("id") == new_id for r in self.cfg["operation_rules"]):
            messagebox.showwarning("提示", "规则编号已存在", parent=self)
            return
        self.cfg["operation_rules"][idx] = dlg.result
        self.save_callback()
        self.current_category = dlg.result.get("output_op")
        self._refresh_categories()

    def _delete_rule(self):
        idx = self._selected_rule_index()
        if idx is None:
            return
        if messagebox.askyesno("确认删除", "确定删除所选规则？", parent=self):
            self.cfg["operation_rules"].pop(idx)
            self.save_callback()
            self._refresh_rules()

    def _toggle_rule(self):
        idx = self._selected_rule_index()
        if idx is None:
            return
        r = self.cfg["operation_rules"][idx]
        r["enabled"] = not bool(r.get("enabled"))
        if r["enabled"] and r.get("invalid_reason"):
            r.pop("invalid_reason", None)
        self.save_callback()
        self._refresh_rules()


class App(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title(APP_NAME)
        self.resizable(True, True)
        self.configure(bg=BG)
        self._cfg = load_config()
        self._debounce_id = None
        self._window_save_id = None
        self._build_ui()
        self._apply_config()
        self.bind("<Configure>", self._on_window_configure)

    def _save_cfg(self):
        self._cfg = save_config(self._cfg)

    def _set_setting(self, key, value):
        self._cfg["settings"][key] = value
        self._save_cfg()

    def _apply_config(self):
        settings = self._cfg["settings"]
        self._dir_var.set(settings.get("out_dir") or DEFAULT_OUT_DIR)
        self._auto_open_var.set(bool(settings.get("auto_open", True)))
        try:
            self.geometry(
                settings.get("window_geometry")
                or DEFAULT_SETTINGS["window_geometry"])
        except Exception:
            self.geometry(DEFAULT_SETTINGS["window_geometry"])
        if settings.get("window_state") == "zoomed":
            self.after(100, lambda: self.state("zoomed"))
        self._update_preview()

    def _build_ui(self):
        bar = tk.Frame(self, bg=DARK, height=52)
        bar.pack(fill="x")
        bar.pack_propagate(False)
        tk.Label(bar,
                 text="TradeCast 交易记录转换助手",
                 bg=DARK,
                 fg="#FFFFFF",
                 font=("微软雅黑", 13, "bold")).pack(side="left", padx=20, pady=14)
        tk.Label(bar,
                 text="交易记录格式转换  ·  TXT / CSV / XLS / XLSX",
                 bg=DARK,
                 fg="#999999",
                 font=("微软雅黑", 9)).pack(side="left", padx=4)

        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=20, pady=16)

        left = tk.Frame(body, bg=BG)
        left.pack(side="left", fill="both", expand=True)
        label_row = tk.Frame(left, bg=BG)
        label_row.pack(fill="x", pady=(0, 6))
        tk.Label(label_row,
                 text="原始记录",
                 bg=BG,
                 fg=DARK,
                 font=("微软雅黑", 10, "bold")).pack(side="left")
        tk.Label(label_row,
                 text="粘贴内容或从文件读取，支持随时手动修改",
                 bg=BG,
                 fg=MID,
                 font=("微软雅黑", 8)).pack(side="left", padx=8)

        text_card = self._card(left)
        text_card.pack(fill="both", expand=True)
        tf = tk.Frame(text_card, bg=PANEL)
        tf.pack(fill="both", expand=True, padx=1, pady=1)

        sb_y = tk.Scrollbar(tf, orient="vertical")
        sb_y.pack(side="right", fill="y")
        sb_x = tk.Scrollbar(tf, orient="horizontal")
        sb_x.pack(side="bottom", fill="x")
        self._text_input = tk.Text(tf,
                                   yscrollcommand=sb_y.set,
                                   xscrollcommand=sb_x.set,
                                   bg="#FAFAFA",
                                   fg=DARK,
                                   insertbackground=DARK,
                                   font=("Consolas", 9),
                                   relief="flat",
                                   borderwidth=0,
                                   highlightthickness=0,
                                   wrap="none")
        self._text_input.pack(side="left", fill="both", expand=True)
        sb_y.config(command=self._text_input.yview)
        sb_x.config(command=self._text_input.xview)
        self._text_input.bind("<<Modified>>", self._on_text_modified)

        btn_row = tk.Frame(left, bg=BG)
        btn_row.pack(fill="x", pady=(8, 0))
        self._btn(btn_row, "清空内容", self._clear_text, True).pack(side="left",
                                                                padx=(0, 8))
        self._btn(btn_row, "从文件读取", self._load_file).pack(side="left")
        self._detect_label = tk.Label(btn_row,
                                      text="",
                                      bg=BG,
                                      fg=MID,
                                      font=("微软雅黑", 8))
        self._detect_label.pack(side="left", padx=12)

        right = tk.Frame(body, bg=BG, width=280)
        right.pack(side="right", fill="y", padx=(16, 0))
        right.pack_propagate(False)

        dir_card = self._card(right)
        dir_card.pack(fill="x")
        tk.Label(dir_card, text="输出目录", bg=PANEL, fg=MID,
                 font=("微软雅黑", 9)).pack(anchor="w", padx=14, pady=(12, 4))
        dir_row = tk.Frame(dir_card, bg=PANEL)
        dir_row.pack(fill="x", padx=14, pady=(0, 12))
        self._dir_var = tk.StringVar()
        tk.Entry(dir_row,
                 textvariable=self._dir_var,
                 bg="#FAFAFA",
                 fg=DARK,
                 font=("微软雅黑", 9),
                 relief="flat",
                 borderwidth=0,
                 highlightthickness=1,
                 highlightbackground=BORDER).pack(side="left",
                                                  fill="x",
                                                  expand=True,
                                                  ipady=5)
        self._btn(dir_row, "浏览", self._browse_dir, True).pack(side="left",
                                                              padx=(6, 0))
        self._dir_var.trace_add("write", self._on_dir_changed)

        name_card = self._card(right)
        name_card.pack(fill="x", pady=(10, 0))
        tk.Label(name_card,
                 text="文件名（自动生成，可手动修改）",
                 bg=PANEL,
                 fg=MID,
                 font=("微软雅黑", 9)).pack(anchor="w", padx=14, pady=(12, 4))
        self._name_var = tk.StringVar(value="交易记录导出")
        tk.Entry(name_card,
                 textvariable=self._name_var,
                 bg="#FAFAFA",
                 fg=DARK,
                 font=("微软雅黑", 9),
                 relief="flat",
                 borderwidth=0,
                 highlightthickness=1,
                 highlightbackground=BORDER).pack(fill="x",
                                                  padx=14,
                                                  pady=(0, 4),
                                                  ipady=5)

        ext_row = tk.Frame(name_card, bg=PANEL)
        ext_row.pack(fill="x", padx=14, pady=(0, 10))
        self._auto_open_var = tk.BooleanVar(value=True)
        tk.Checkbutton(ext_row,
                       text="自动打开成果文件",
                       variable=self._auto_open_var,
                       bg=PANEL,
                       fg=MID,
                       font=("微软雅黑", 8),
                       activebackground=PANEL,
                       command=self._on_auto_open_changed).pack(side="left")
        tk.Label(ext_row,
                 text=".xlsx",
                 bg=PANEL,
                 fg="#AAAAAA",
                 font=("微软雅黑", 8)).pack(side="right")

        preview_card = self._card(right)
        preview_card.pack(fill="x", pady=(10, 0))
        tk.Label(preview_card,
                 text="完整路径预览",
                 bg=PANEL,
                 fg=MID,
                 font=("微软雅黑", 9)).pack(anchor="w", padx=14, pady=(10, 4))
        self._preview_label = tk.Label(preview_card,
                                       text="",
                                       bg=PANEL,
                                       fg="#888888",
                                       font=("微软雅黑", 8),
                                       wraplength=240,
                                       justify="left")
        self._preview_label.pack(anchor="w", padx=14, pady=(0, 10))
        self._name_var.trace_add("write", lambda *_: self._update_preview())

        self._btn(right,
                  "开始转换  →",
                  self._run,
                  font=("微软雅黑", 11, "bold"),
                  pady=14).pack(fill="x", pady=(14, 0))
        self._btn(right, "规则维护", self._open_rule_manager,
                  True).pack(fill="x", pady=(10, 0))

        tk.Label(right, text="运行日志", bg=BG, fg=MID,
                 font=("微软雅黑", 9)).pack(anchor="w", pady=(16, 4))
        log_card = self._card(right)
        log_card.pack(fill="both", expand=True)
        sb = tk.Scrollbar(log_card)
        sb.pack(side="right", fill="y")
        self._log = tk.Text(log_card,
                            yscrollcommand=sb.set,
                            bg=PANEL,
                            fg=MID,
                            font=("微软雅黑", 8),
                            relief="flat",
                            borderwidth=0,
                            highlightthickness=0,
                            state="disabled",
                            wrap="word")
        self._log.pack(fill="both", expand=True, padx=8, pady=8)
        sb.config(command=self._log.yview)

        status = tk.Frame(self, bg="#E8E8E8", height=28)
        status.pack(fill="x", side="bottom")
        status.pack_propagate(False)
        self._status_var = tk.StringVar(value="就绪")
        tk.Label(status,
                 textvariable=self._status_var,
                 bg="#E8E8E8",
                 fg=MID,
                 font=("微软雅黑", 8)).pack(side="left", padx=14, pady=6)

    def _card(self, parent):
        return tk.Frame(parent,
                        bg=PANEL,
                        relief="flat",
                        highlightbackground=BORDER,
                        highlightthickness=1)

    def _btn(self, parent, text, cmd, ghost=False, font=("微软雅黑", 9), pady=4):
        box = tk.Frame(parent, bg=SHADOW)
        btn = tk.Button(
            box,
            text=text,
            command=cmd,
            font=font,
            relief="flat",
            cursor="hand2",
            padx=10,
            pady=pady,
            bd=0,
            highlightthickness=1,
            highlightbackground="#FFFFFF" if ghost else "#6B7680",
            bg=BTN_GHOST if ghost else BTN_DARK,
            fg="#4A4A4A" if ghost else "#FFFFFF",
            activebackground=BTN_GHOST_ACTIVE if ghost else BTN_DARK_ACTIVE,
            activeforeground="#333333" if ghost else "#FFFFFF")
        btn.pack(fill="both", expand=True, padx=(0, 2), pady=(0, 2))
        return box

    def _on_window_configure(self, event=None):
        if event is not None and event.widget is not self:
            return
        if self._window_save_id:
            self.after_cancel(self._window_save_id)
        self._window_save_id = self.after(700, self._save_window_state)

    def _save_window_state(self):
        try:
            state = self.state()
            self._cfg["settings"]["window_state"] = state
            if state == "normal":
                self._cfg["settings"]["window_geometry"] = self.geometry()
            self._save_cfg()
        except Exception:
            pass

    def _on_dir_changed(self, *_):
        self._update_preview()
        self._set_setting("out_dir",
                          clean_path(self._dir_var.get()) or DEFAULT_OUT_DIR)

    def _on_auto_open_changed(self):
        self._set_setting("auto_open", bool(self._auto_open_var.get()))

    def _on_text_modified(self, event=None):
        self._text_input.edit_modified(False)
        if self._debounce_id:
            self.after_cancel(self._debounce_id)
        self._debounce_id = self.after(600, self._auto_detect)

    def _auto_detect(self):
        text = self._text_input.get("1.0", "end").strip()
        if not text:
            self._detect_label.config(text="")
            self._name_var.set("交易记录导出")
            return

        blocks = parse_record_blocks(text)
        if not blocks:
            self._detect_label.config(text="⚠ 未识别到有效记录", fg="#CC6600")
            return

        type1_count = sum(len(b["records"]) for b in blocks if b["rtype"] == 1)
        type2_count = sum(len(b["records"]) for b in blocks if b["rtype"] == 2)

        parts = []
        if type1_count:
            parts.append(f"成交记录 {type1_count} 行")
        if type2_count:
            parts.append(f"资金流水 {type2_count} 行")

        self._detect_label.config(text="✓ 已识别：" + "；".join(parts),
                                  fg="#2A7A2A")
        self._name_var.set(infer_filename(text))

    def _update_preview(self):
        d = self._dir_var.get().strip() or DEFAULT_OUT_DIR
        n = self._name_var.get().strip() or "交易记录导出"
        self._preview_label.config(text=os.path.join(d, n + ".xlsx"))

    def _clear_text(self):
        self._text_input.delete("1.0", "end")
        self._detect_label.config(text="")
        self._name_var.set("交易记录导出")

    def _load_file(self):
        init_dir = self._cfg["settings"].get(
            "last_open_dir") or DEFAULT_OPEN_DIR
        paths = filedialog.askopenfilenames(title="选择原始记录文件",
                                            initialdir=init_dir,
                                            filetypes=[
                                                ("支持的文件",
                                                 "*.txt *.csv *.xls *.xlsx"),
                                                ("所有文件", "*.*")
                                            ])
        if not paths:
            return

        self._set_setting("last_open_dir",
                          clean_path(os.path.dirname(paths[-1])))

        all_text = []
        for path in paths:
            ext = os.path.splitext(path)[1].lower()
            try:
                if ext in (".xls", ".xlsx"):
                    df = pd.read_excel(path, dtype=str, header=None).fillna("")
                    all_text.append("\n".join("\t".join(str(v) for v in row)
                                              for _, row in df.iterrows()))
                else:
                    with open(path, "rb") as f:
                        raw = f.read()
                    all_text.append(
                        raw.decode(detect_encoding(raw), errors="replace"))
            except Exception as e:
                messagebox.showerror("读取失败", f"{os.path.basename(path)}\n{e}")
                return

        self._text_input.delete("1.0", "end")
        self._text_input.insert("1.0", "\n\n".join(all_text))
        self._status_var.set(f"已载入 {len(paths)} 个文件")

    def _browse_dir(self):
        d = filedialog.askdirectory(title="选择输出目录",
                                    initialdir=self._dir_var.get().strip()
                                    or DEFAULT_OUT_DIR)
        if d:
            self._dir_var.set(d)

    def _get_out_path(self):
        return os.path.join(self._dir_var.get().strip() or DEFAULT_OUT_DIR,
                            (self._name_var.get().strip() or "交易记录导出") +
                            ".xlsx")

    def _log_write(self, msg):
        self._log.config(state="normal")
        self._log.insert("end", msg + "\n")
        self._log.see("end")
        self._log.config(state="disabled")

    @staticmethod
    def _is_file_open(path):
        try:
            with open(path, "a+b"):
                return False
        except IOError:
            return True

    def _pick_saveas_path(self, current_path):
        return filedialog.asksaveasfilename(
            initialdir=os.path.dirname(current_path),
            initialfile=os.path.basename(current_path),
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")])

    def _resolve_output_path(self, path):
        if os.path.exists(path):
            choice = messagebox.askyesnocancel(
                "文件已存在", f"「{os.path.basename(path)}」已存在。\n\n是=覆盖　否=另存为　取消=终止")
            if choice is None:
                return None
            if choice is False:
                new_path = self._pick_saveas_path(path)
                return self._resolve_output_path(
                    new_path) if new_path else None

        while self._is_file_open(path):
            retry = messagebox.askyesnocancel(
                "文件被占用",
                f"「{os.path.basename(path)}」正被其他程序打开，无法写入。\n\n请关闭该文件后点击「是」重试\n否=另存为　取消=终止"
            )
            if retry is None:
                return None
            if retry is False:
                new_path = self._pick_saveas_path(path)
                return self._resolve_output_path(
                    new_path) if new_path else None
        return path

    def _open_rule_manager(self):
        RuleManager(self, self._cfg, self._save_cfg)

    def _run(self):
        text = self._text_input.get("1.0", "end").strip()
        if not text:
            messagebox.showwarning("提示", "请先输入或载入原始记录内容")
            return

        out_path = self._resolve_output_path(self._get_out_path())
        if out_path is None:
            self._status_var.set("已取消")
            return

        self._name_var.set(os.path.splitext(os.path.basename(out_path))[0])
        self._status_var.set("处理中…")
        self.update_idletasks()

        ok, msg = process_text(text, out_path, self._cfg)
        self._log_write(msg)
        self._status_var.set("✅ 完成" if ok else "❌ 出错")
        (messagebox.showinfo if ok else messagebox.showerror)(
            "完成" if ok else "错误", msg)

        if ok and self._auto_open_var.get():
            os.startfile(out_path)


if __name__ == "__main__":
    App().mainloop()
